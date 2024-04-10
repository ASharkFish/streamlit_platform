from tqdm import tqdm
from time import sleep
import pandas as pd 
import numpy as np
from time import sleep
import warnings
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
import datetime

import 细分市场品牌品类检查V1_2
import 规格检查v1_7
import 月报对比V3_2
import 字段对应V1_4
import 字段有效性及组合有效性V3_9
import 字段值空值检查V2_8
import 一对多新版规则V2_2
import 价格逻辑检查V2_7
import 店铺类型异常_新版规则V2_12


import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))
from main import sql_connect
from main import common


pd.set_option('display.max_rows', None) 


warnings.filterwarnings("ignore")
pd.set_option('display.max_rows', 100)

def sql_whole():
    sql=f'''
        select CAST ( 客户 AS nvarchar ), CAST ( 品类 AS nvarchar ),CAST ( 客户品类 AS nvarchar ),CAST ( 数据库名 AS nvarchar ) from QC.dbo.客户品类对照表\
        where 判断类型='分类'
    '''
    print(sql)
    database='QC'
    column=['客户','品类','客户品类','数据库']
    df=sql_connect.select('192.168.0.15',database,'sqlsever',sql,column)
    return df

def score(df):
    rows, columns = df.shape
    if rows <= 10:
        return [10 - rows,rows]
    else:
        return [0,rows]
    

def score_2(df):
    rows, columns = df.shape
    if rows <= 20:
        return [20 - rows,rows]
    else:
        return [0,rows]
    


def score_s(df1,df2,df3):
    row1, column1 = df1.shape
    row2, column2 = df2.shape
    row3, column3 = df3.shape
    rows=row1+row2+row3
    if rows <= 10:
        return [10 - rows,rows]
    else:
        return [0,rows]



def main(kehu_list,time1,time2):
    df=sql_whole()
    
    
    python_path,file_total=common.path()

    result_path=f'{file_total}/result'
    template_path=f'{file_total}/template'



    time_lm=(pd.Timestamp(time1[:4]+'-'+time1[-2:])+pd.DateOffset(n=-1, months=1)).strftime("%Y-%m").replace('-','')

    customer_total=list(df['客户品类'].unique())
    customer=kehu_list[0]
    database= df.loc[df['客户品类'] == customer, '数据库'].to_string(index=False)
    category=df.loc[df['客户品类'] == customer, '品类'].to_string(index=False)
    print(database,category)
    error=[]
    name_list=['细分市场品牌品类检查','店铺类型异常','规格检查','字段值一对多','字段有效性,组合有效性及分组判断',
               '字段值空值检查','月报对比','字段对应','价格逻辑检查']
    score_list=[]
    shape_list=[]
    
    current_time = datetime.datetime.now()
    formatted_time = current_time.strftime("%Y-%m-%d %H_%M_%S")

    template_path=f'{template_path}/数据质检报告.xlsx'
    filename = f'{result_path}/数据质检报告{customer}{time1}-{time2}-{formatted_time}.xlsx'
    wb = openpyxl.load_workbook(template_path)





    print('细分市场品牌品类检查')
    try:
        df_segmenting=细分市场品牌品类检查V1_2.main(customer,time1,time2)
        score_list.append(score_2(df_segmenting)[0])
        shape_list.append(score_2(df_segmenting)[1])
        error.append('已关联')
        worksheet = wb.create_sheet(title='细分市场品牌品类检查')
        for row in dataframe_to_rows(df_segmenting, index=False, header=True):
            worksheet.append(row)
    except:
        error.append('细分市场品牌品类检查部分无数据或异常')
        score_list.append(20)
        shape_list.append(0)



    print('店铺类型异常')
    try:
        store_type=店铺类型异常_新版规则V2_12.main(customer,time1,time2)
        score_list.append(score(store_type)[0])
        shape_list.append(score(store_type)[1])
        error.append('已关联')
        worksheet = wb.create_sheet(title='店铺类型异常')
        for row in dataframe_to_rows(store_type, index=False, header=True):
            worksheet.append(row)
    except:
        error.append('店铺类型异常部分无数据或异常')
        score_list.append(10)
        shape_list.append(0)



    print('规格检查')
    try:
        specifications_result=规格检查v1_7.main(customer,time1,time2)
        score_list.append(score(specifications_result)[0])
        shape_list.append(score(specifications_result)[1])
        error.append('已关联')
        worksheet = wb.create_sheet(title='规格检查')
        for row in dataframe_to_rows(specifications_result, index=False, header=True):
            worksheet.append(row)
    except:
        error.append('规格检查部分无数据或异常')
        score_list.append(10)
        shape_list.append(0)



    print('字段值一对多')
    try:
        one_to_many=一对多新版规则V2_2.main(customer,time1,time2)
        score_list.append(score(one_to_many)[0])
        shape_list.append(score(one_to_many)[1])
        error.append('已关联')
        worksheet = wb.create_sheet(title='字段值一对多')
        for row in dataframe_to_rows(one_to_many, index=False, header=True):
            worksheet.append(row)

    except:
        error.append('字段值一对多部分无数据或异常')
        score_list.append(10)
        shape_list.append(0)



    print('字段有效性,组合有效性及分组判断')
    try:
        effectiveness,combination_effectiveness,grouping_judgment=字段有效性及组合有效性V3_9.main(customer)
        score_list.append(score_s(effectiveness,combination_effectiveness,grouping_judgment)[0])

        shape_list.append(score_s(effectiveness,combination_effectiveness,grouping_judgment)[1])

        error.append('已关联')

        worksheet = wb.create_sheet(title='字段有效性')
        for row in dataframe_to_rows(effectiveness, index=False, header=True):
            worksheet.append(row)

        worksheet = wb.create_sheet(title='字段组合有效性')
        for row in dataframe_to_rows(combination_effectiveness, index=False, header=True):
            worksheet.append(row)

        worksheet = wb.create_sheet(title='分组判断')
        for row in dataframe_to_rows(grouping_judgment, index=False, header=True):
            worksheet.append(row)

    except:
        error.append('字段有效性,组合有效性及分组判断部分无数据或异常')

        score_list.append(10)
        
        shape_list.append(0)



    print('字段值空值检查')
    try:
        null_value=字段值空值检查V2_8.main(customer,time1,time2)
        score_list.append(score(null_value)[0])
        shape_list.append(score(null_value)[1])
        error.append('已关联')
        worksheet = wb.create_sheet(title='字段值空值检查')
        for row in dataframe_to_rows(null_value, index=False, header=True):
            worksheet.append(row)

    except:
        error.append('字段值空值检查部分无数据或异常')
        score_list.append(10)
        shape_list.append(0)



    print('月报对比')
    try:
        monthly_report=月报对比V3_2.main(customer,time_lm)
        score_list.append(score(monthly_report)[0])
        shape_list.append(score(monthly_report)[1])
        error.append('已关联')
        worksheet = wb.create_sheet(title='月报对比')
        for row in dataframe_to_rows(monthly_report, index=False, header=True):
            worksheet.append(row)

    except:
        error.append('月报对比部分无数据或异常')
        score_list.append(10)
        shape_list.append(0)


    print('字段对应')
    try: 
        field_correspondence=字段对应V1_4.main(customer,time1,time2)
        score_list.append(score(field_correspondence)[0])
        shape_list.append(score(field_correspondence)[1])
        error.append('已关联')
        worksheet = wb.create_sheet(title='字段对应')
        for row in dataframe_to_rows(field_correspondence, index=False, header=True):
            worksheet.append(row)

    except:
        error.append('字段对应部分无数据或异常')
        score_list.append(10)
        shape_list.append(0)


    print('价格逻辑检查')
    try: 
        price_logic=价格逻辑检查V2_7.main(customer,time1,time2)
        score_list.append(score(price_logic)[0])
        shape_list.append(score(price_logic)[1])
        error.append('已关联')
        worksheet = wb.create_sheet(title='价格逻辑检查')
        for row in dataframe_to_rows(price_logic, index=False, header=True):
            worksheet.append(row)

    except:
        error.append('价格逻辑检查部分无数据或异常')
        score_list.append(10)
        shape_list.append(0)

    

    score_total=sum(score_list)
    name_list.append('总得分')
    score_list.append(score_total)
    error.append('无')
    shape_list.append(np.nan)

    df_output=pd.DataFrame({
        'name':name_list,
        'score':score_list,
        'notes':error,
        'error_num':shape_list
    })
    order=['细分市场品牌品类检查','店铺类型异常','规格检查','字段值一对多','字段有效性,组合有效性及分组判断',
               '字段值空值检查','月报对比','字段对应','价格逻辑检查','总得分']

    df_output['name'] = pd.Categorical(df_output['name'], categories=order, ordered=True)
    df_output = df_output.sort_values('name')
    df_output=df_output.set_index('name')


    print(error)


    start_col=6                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                 
    start_row=4

    ws = wb['数据库质检报告']
    ws['B1'] =f'开始时间:{time1},结束时间:{time2}'

    if customer in customer_total:
        ws['B15'] =f'已关联客户品类名:{customer}'
    else:
        print(f'无客户名:{customer}')
        ws['B15'] =f'无客户品类名:{customer}'
        df_output['score']=0
        score_total=0



    for r in dataframe_to_rows(df_output, index=False, header=False):
        for col_idx, cell_value in enumerate(r, start_col):
            ws.cell(row=start_row, column=col_idx, value=cell_value) 
        start_row += 1
    fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    if score_total <= 60:
        cell = ws['F13']
        cell.fill = fill

    ws['B15'] =f'数据库运行时间:{formatted_time}'

    wb.save(filename)

    print('输出完成')





if __name__ == "__main__":

    received_arguments = sys.argv[1:]
    kehu_list=received_arguments[0].split(',')
    time_begin=received_arguments[1]
    time_end=received_arguments[2]
    main(kehu_list,time_begin,time_end)