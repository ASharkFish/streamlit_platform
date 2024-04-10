import pandas as pd
from tqdm import tqdm
from functools import reduce
import numpy as np
from datetime import datetime
import openpyxl
import warnings
from openpyxl.utils.dataframe import dataframe_to_rows
from dateutil.relativedelta import relativedelta

import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))
from main import sql_connect
from main import common

warnings.filterwarnings("ignore")
#sql查询
def sql_select(brand_list):
    sql = "select 月份, cast(BRAND as nvarchar),子品类,cast(CHANNEL as nvarchar),cast(SEGMENT as nvarchar),\
            [CHANNEL_TYPE_NEW],cast(产品名称 as nvarchar(300)),\
            [Value(RMB,000)],[Volume(KG)] from send_out.dbo.BAISHI_DATA_MAIPIAN_E_NEW \
            where 月份  >= 201701  and BRAND!= '虚拟定制' "
    print(sql)
    column=['时间','brand','子品类','CHANNEL','SEGMENT','CHANNEL_TYPE_NEW','产品名称','Value','Volume']
    df=sql_connect.select('192.168.0.15','send_out','sqlsever',sql,column)
    df['Volume']=df['Volume']/1000

    df['brand']=df['brand'].apply(lambda x: x if x in brand_list else 'Others')

    return df 


#ytd计算函数
def generate_ytd_list(input_time):
    date = datetime.strptime(input_time, '%Y-%m')
    
    year = date.year
    month = date.month
    
    ytd_list = []
    for i in range(1, month+1):
        month_str = str(i).rjust(2, '0')
        ytd_month = f'{year}{month_str}'
        
        ytd_list.append(ytd_month)
    
    return ytd_list
#FY计算函数
def generate_yearly_list(input_time):
    date = datetime.strptime(input_time, '%Y-%m')
    
    year = date.year

    yearly_list = []

    for i in range(1, 13):
        monthly_str = str(i).rjust(2, '0')
        yearly_list.append(f'{year}{monthly_str}')
    
    return yearly_list
#pm时间计算函数
def previous_n_months(input_time,key):
    date = datetime.strptime(input_time, '%Y-%m')
    result_current = []
    result_prev_1=[]
    result_prev_3 = []
    result_prev_6 = []
    result_prev_12 = []

    for i in range(12):
        if i == 1:
            p_1_month = date - relativedelta(months=i)
            result_prev_1.append(p_1_month.strftime('%Y%m'))

        if i < 3:
            p_3_month = date - relativedelta(months=i)
            result_prev_3.append(p_3_month.strftime('%Y%m'))

        if i < 6:
            p_6_month = date - relativedelta(months=i)
            result_prev_6.append(p_6_month.strftime('%Y%m'))

        if i < 12:
            p_12_month = date - relativedelta(months=i)
            result_prev_12.append(p_12_month.strftime('%Y%m'))

    result_current=[input_time.replace('-','')]
    column_name=[f'P1M {key}',f'P12M {key}',f'P3M {key}',f'P6M {key}']

    return [result_current,result_prev_12, result_prev_3, result_prev_6],column_name,result_prev_1
#24月计算函数
def past_24_months(input_time):
    result = []
    current_date = datetime.strptime(input_time, '%Y%m')
    for i in range(25):
        past_date = current_date - relativedelta(months=i)
        result.append(past_date.strftime('%Y%m'))
    return result
#201701至今计算函数
def time_17(time):
    list=[]
    m=0
    input='2017-01'
    output=''
    while output < time :
        output=(pd.Timestamp(input)+pd.DateOffset(n=m, months=1)).strftime("%Y-%m")
        list.append(output.replace('-',''))
        m+=1
    return list 
#201801至今计算函数
def time_18(time):
    list=[]
    m=0
    input='2018-01'
    output=''
    while output < time :
        output=(pd.Timestamp(input)+pd.DateOffset(n=m, months=1)).strftime("%Y-%m")
        list.append(output.replace('-',''))
        m+=1
    return list 
#时间及对应列名输出部分
def time_total(time):
    time_ty=time[:4]+'-'+time[-2:]
    num=int(time[:4])-2018
    time_list=[]
    for i in range(num):
        time_list.append((pd.Timestamp(time_ty)+pd.DateOffset(n=-i, years=1)).strftime("%Y-%m"))
    
    ytd_list=[]
    year_list=[]
    column_ytd=[]
    column_year=[]
    for date in time_list:
        year=date[:4]
        ytd_list.append(generate_ytd_list(date))
        column_ytd.append(f'YTD {year}')
        if date != time_ty:
            year_list.append(generate_yearly_list(date))
            column_year.append(f'FY {year}')
    ytd_list.reverse()
    year_list.reverse()
    column_ytd.reverse()
    column_year.reverse()

    p_17_month=time_17(time_ty)
    p_18_month=time_18(time_ty)

    pm_ty,column_pm_ty,p1m_lm_ty=previous_n_months(time_list[0],'TY')
    pm_ly,column_pm_ly,p1m_lm_ly=previous_n_months(time_list[1],'LY')
    p_24_month=past_24_months(time)
    p_24_month.reverse()

    column_pm_ty.append('P1M LM')
    pm_ty.append(p1m_lm_ty)
    return ytd_list,year_list,pm_ty,pm_ly,p_24_month,p_17_month,p_18_month,column_ytd,column_year,column_pm_ty,column_pm_ly



#sale同环比计算、列排序
def sale_count(df_input,column_pm_ty,column_pm_ly,column_ytd):
    column_list=[]
    df=df_input[['platform','Category/subcategory']+column_pm_ty+column_pm_ly+column_ytd].copy()
    df_input_columns_list = df_input.columns.tolist()
    other_list = [x for x in df_input_columns_list if x not in column_pm_ty+column_pm_ly+column_ytd]

    df_other=df_input[other_list].copy()

    df['YTD YA']=df[column_ytd[-1]]/df[column_ytd[-2]]-1
    for num in range(len(column_pm_ly)):
        column1=column_pm_ty[num]
        column2=column_pm_ly[num]
        if column1=='P1M TY':
            column_output='P1M TY YA'
            column_output_s='P1M TY pp'
            column_s='P1M LM'

            df[column_output]=df[column1]/df[column2]-1
            df[column_output_s]=df[column1]/df[column_s]-1

            column_list.append(column1)
            column_list.append(column_output_s)
            column_list.append(column_output)

        else:
            column_output=f'{column1} YA'
            df[column_output]=df[column1]/df[column2]-1
            
            column_list.append(column2)
            column_list.append(column1)
            column_list.append(column_output)
    df=df[['platform','Category/subcategory']+column_list+column_ytd+['YTD YA']]

    result=pd.merge(df,df_other,how='left',on=['platform','Category/subcategory']).reset_index(drop=True)
    return result 


#Sshare 同环比计算、列排序
def share_count(df_input,column_pm_ty,column_pm_ly,column_ytd):
    column_list=[]
    df=df_input[['platform','Category/subcategory']+column_pm_ty+column_pm_ly+column_ytd].copy()
    df_input_columns_list = df_input.columns.tolist()
    other_list = [x for x in df_input_columns_list if x not in column_pm_ty+column_pm_ly+column_ytd]

    df_other=df_input[other_list].copy()

    df['YTD YA']=df[column_ytd[-1]]-df[column_ytd[-2]]
    for num in range(len(column_pm_ly)):
        column1=column_pm_ty[num]
        column2=column_pm_ly[num]
        if column1=='P1M TY':
            column_output='P1M TY YA'
            column_output_s='P1M TY pp'
            column_s='P1M LM'

            df[column_output]=df[column1]-df[column2]
            df[column_output_s]=df[column1]-df[column_s]

            column_list.append(column1)
            column_list.append(column_output_s)
            column_list.append(column_output)

        else:
            column_output=f'{column1} YA'
            df[column_output]=df[column1]-df[column2]
            
            column_list.append(column2)
            column_list.append(column1)
            column_list.append(column_output)
    df=df[['platform','Category/subcategory']+column_list+column_ytd+['YTD YA']]

    result=pd.merge(df,df_other,how='left',on=['platform','Category/subcategory']).reset_index(drop=True)
    return result 


#多df merge聚合函数
def list_merge(df_list,on,how):
    merged_df = reduce(lambda left, right: pd.merge(left, right, on=on, how=how), df_list)
    return merged_df

#主体计算部分
def total_count(df_total,time_list,column_list,count_word,key_word,total_word,platform_name):
    df_test=df_total[[key_word]].drop_duplicates().reset_index(drop=True)
    if time_list!=column_list:
        sum_list=[total_word]
        for num in range(len(time_list)):
            column_name=column_list[num]

            df_input=df_total[df_total['时间'].isin(time_list[num])][[key_word,count_word]].groupby([key_word],as_index=False).sum()

            sum_list.append(df_input[count_word].sum())
            df_output=df_input.rename(columns={count_word:f'{column_name}'})
            df_test=pd.merge(df_test,df_output,how='left',on=[key_word])
        df_test.loc[len(df_test)] = sum_list
    else:  
        sum_list=[]

        for time in time_list:
            df_input=df_total[df_total['时间']==time]
            if df_input.empty:
                sum_list.append(np.nan)
            else:
                sum_list.append(df_input[count_word].sum())

        dic={
            key_word:total_word,
            '时间':time_list,
            count_word:sum_list
        }
        df_linshi1=pd.DataFrame(dic)

        df_input=df_total[['时间',key_word,count_word]].groupby(['时间',key_word],as_index=False).sum()
        df_input=pd.concat([df_input,df_linshi1]).drop_duplicates().reset_index(drop=True)
        if total_word in ['B2C-Tmall','B2C-standalone B2C']:
            df_test=df_linshi1.pivot(index=[key_word],columns='时间',values=count_word).rename_axis(columns=None).reset_index()
            df_test=df_test[[key_word]+time_list].reset_index(drop=True)
        else:
            df_test=df_input.pivot(index=[key_word],columns='时间',values=count_word).rename_axis(columns=None).reset_index()

            df_test=df_test[[key_word]+time_list].reset_index(drop=True)
    
    df_test['platform']=platform_name
    return df_test









#brand各时间部分聚合拼接
def brand_polymerization(df_total,df_channel_value,ytd_list,year_list,pm_ty,pm_ly,column_pm_ty,column_pm_ly,p_24_month,key_list,platform_name):
    
    count_word=key_list[0]
    key_word=key_list[1]
    total_word=key_list[2]

    ytd_test=total_count(df_total,ytd_list,column_ytd,count_word,key_word,total_word,platform_name)
    year_test=total_count(df_total,year_list,column_year,count_word,key_word,total_word,platform_name)
    pm_test_ty=total_count(df_total,pm_ty,column_pm_ty,count_word,key_word,total_word,platform_name)
    pm_test_ly=total_count(df_total,pm_ly,column_pm_ly,count_word,key_word,total_word,platform_name)
    p24_month_test=total_count(df_total,p_24_month,p_24_month,count_word,key_word,total_word,platform_name)
    channel_value=list_merge([df_channel_value,pm_test_ty,pm_test_ly,ytd_test,year_test,p24_month_test],['platform',key_word],'left')
    
    channel_value=channel_value.drop_duplicates().reset_index(drop=True)
    channel_value=channel_value.rename(columns={key_word:'Category/subcategory'})
    return channel_value


#brand二次聚合
def brand_polymerization_2(df_total,dic,count_word,rule):
    output_list=[]
    platform_name=dic.loc[0,'value_word']
    for num in range(len(dic)):
        screen_word=dic.loc[num,'screen_word']
        value_word=dic.loc[num,'value_word']
        column_word=dic.loc[num,'column_word']
        column_value=dic.loc[num,'column_value']

        if value_word!='Total':
            df_input=df_total[df_total[screen_word]==value_word].reset_index(drop=True)
        else:
            df_input=df_total.copy()
        df_channel=pd.DataFrame({   'platform':platform_name,
                                    column_word:column_value 
                                })
        key_list=[count_word,column_word,value_word]
        result=brand_polymerization(df_input,df_channel,ytd_list,year_list,pm_ty,pm_ly,column_pm_ty,column_pm_ly,p_24_month,key_list,platform_name)
        output_list.append(result)
    df_result=pd.concat(output_list).reset_index(drop=True)

    if rule=='share':

        df_result.set_index(['platform','Category/subcategory'],inplace=True)
        rule_values = df_result.loc[(platform_name,platform_name)]
        df_result = df_result.div(rule_values)
        df_result.reset_index(inplace=True)

    return df_result


#brand第一部分聚合
def polymerization_total(df_total,count_word,rule,name):
    #平台==total
    if name=='Total Cereal':
        dic_total=pd.DataFrame({
            'screen_word':['CHANNEL','SEGMENT','SEGMENT','SEGMENT'],
            'value_word':['Total','冷食','热食','三合一'],
            'column_word':['CHANNEL','子品类','子品类','子品类'],
            'column_value':[['Total','B2C-Tmall','B2C-standalone B2C','B2C-PDD','抖音'],['冷食','Base Granola','Cold Others','Fancy Granola'],['热食','Base Hot','Hot Others','VA Hot'],['三合一']]
        })
    elif name=='EC':
        dic_total=pd.DataFrame({
            'screen_word':['CHANNEL','SEGMENT','SEGMENT','SEGMENT'],
            'value_word':['Total','冷食','热食','三合一'],
            'column_word':['CHANNEL','子品类','子品类','子品类'],
            'column_value':[['Total','B2C-Tmall','B2C-standalone B2C'],['冷食','Base Granola','Cold Others','Fancy Granola'],['热食','Base Hot','Hot Others','VA Hot'],['三合一']]
        })

    total_value=brand_polymerization_2(df_total,dic_total,count_word,rule)

    #平台==TMALL
    dic_TM=pd.DataFrame({
        'screen_word':['CHANNEL','SEGMENT','SEGMENT','SEGMENT'],
        'value_word':['B2C-Tmall','冷食','热食','三合一'],
        'column_word':['CHANNEL','子品类','子品类','子品类'],
        'column_value':[['B2C-Tmall'],['冷食','Base Granola','Cold Others','Fancy Granola'],['热食','Base Hot','Hot Others','VA Hot'],['三合一']]
    })
    df_input=df_total[df_total['CHANNEL']=='B2C-Tmall']
    TM_value=brand_polymerization_2(df_input,dic_TM,count_word,rule)

    #平台==Standalone
    dic_Standalone=pd.DataFrame({
        'screen_word':['CHANNEL','SEGMENT','SEGMENT','SEGMENT'],
        'value_word':['B2C-standalone B2C','冷食','热食','三合一'],
        'column_word':['CHANNEL','子品类','子品类','子品类'],
        'column_value':[['B2C-standalone B2C'],['冷食','Base Granola','Cold Others','Fancy Granola'],['热食','Base Hot','Hot Others','VA Hot'],['三合一']]
    })
    df_input=df_total[df_total['CHANNEL']=='B2C-standalone B2C']
    Standalone_value=brand_polymerization_2(df_input,dic_Standalone,count_word,rule)


    df_concat=pd.concat([total_value,TM_value,Standalone_value]).reset_index(drop=True)
    if rule=='sale':
        value1=sale_count(df_concat,column_pm_ty,column_pm_ly,column_ytd)
    elif rule=='share':
        value1=share_count(df_concat,column_pm_ty,column_pm_ly,column_ytd)
    return value1


#brand品牌部分聚合
def polymerization_brand(df_total,count_word,rule):
    platform_list=['Total','B2C-Tmall','B2C-standalone B2C']
    df_rule=pd.DataFrame({
        'column_word':['Total','SEGMENT','子品类','子品类','子品类','SEGMENT','子品类','子品类','子品类','SEGMENT',],
        'value_word':['Total','冷食','Base Granola','Cold Others','Fancy Granola','热食','Base Hot','VA Hot','Hot Others','三合一']
    })
    df_list=[]
    for platform in platform_list:
        if platform=='Total':
            df_platform=df_total.copy()
        else:
            df_platform=df_total[df_total['CHANNEL']==platform].reset_index(drop=True)
        for num in range(len(df_rule)):
            column_word=df_rule.loc[num,'column_word']
            value_word=df_rule.loc[num,'value_word']
            if column_word=='Total':
                df_input=df_platform.copy()
            else:
                df_input=df_platform[df_platform[column_word]==value_word]
            dic_brand=pd.DataFrame({
                'screen_word':['brand'],
                'value_word':['Total'],
                'column_word':['brand'],
                'column_value':[brand_list+['Others','Total']]
            })
            df_brand=brand_polymerization_2(df_input,dic_brand,count_word,rule)
            
            if rule=='sale':
                brand_value=sale_count(df_brand,column_pm_ty,column_pm_ly,column_ytd)
            elif rule=='share':
                brand_value=share_count(df_brand,column_pm_ty,column_pm_ly,column_ytd)

            if (platform!='Total')&(column_word=='Total'):
                continue
            else:
                df_list.append(brand_value)
        
    return df_list



#platform 
def platform_polymerization(df_total,df_channel_value,ytd_list,year_list,pm_ty,pm_ly,column_pm_ty,column_pm_ly,p_all_month,key_list,platform_name):
    
    count_word=key_list[0]
    key_word=key_list[1]
    total_word=key_list[2]

    ytd_test=total_count(df_total,ytd_list,column_ytd,count_word,key_word,total_word,platform_name)
    year_test=total_count(df_total,year_list,column_year,count_word,key_word,total_word,platform_name)
    pm_test_ty=total_count(df_total,pm_ty,column_pm_ty,count_word,key_word,total_word,platform_name)
    pm_test_ly=total_count(df_total,pm_ly,column_pm_ly,count_word,key_word,total_word,platform_name)
    p_all_month_test=total_count(df_total,p_all_month,p_all_month,count_word,key_word,total_word,platform_name)

    channel_value=list_merge([df_channel_value,pm_test_ty,pm_test_ly,ytd_test,year_test,p_all_month_test],['platform',key_word],'left')
    
    channel_value=channel_value.drop_duplicates().reset_index(drop=True)
    channel_value=channel_value.rename(columns={key_word:'Category/subcategory'})
    return channel_value


def platform_polymerization_2(df_total,p_all_month,dic,count_word,rule):
    output_list=[]

    platform_name=dic.loc[0,'value_word']
    for num in range(len(dic)):
        screen_word=dic.loc[num,'screen_word']
        value_word=dic.loc[num,'value_word']
        column_word=dic.loc[num,'column_word']
        column_value=dic.loc[num,'column_value']

        if value_word!='Total':
            df_input=df_total[df_total[screen_word]==value_word].reset_index(drop=True)
        else:
            df_input=df_total.copy()
        df_channel=pd.DataFrame({   'platform':platform_name,
                                    column_word:column_value 
                                })
        key_list=[count_word,column_word,value_word]
        result=platform_polymerization(df_input,df_channel,ytd_list,year_list,pm_ty,pm_ly,column_pm_ty,column_pm_ly,p_all_month,key_list,platform_name)
        output_list.append(result)
    df_result=pd.concat(output_list).reset_index(drop=True)

    if rule=='share':

        df_result.set_index(['platform','Category/subcategory'],inplace=True)
        rule_values = df_result.loc[(platform_name,platform_name)]
        df_result = df_result.div(rule_values)
        df_result.reset_index(inplace=True)

    return df_result


def total_channel(df_total,count_word,rule,platform_name,column_list,channel_type_list,p_all_month):
    
    df_input=df_total[df_total['CHANNEL']==platform_name]

    df_list1=[]
    df_list2=[]

    if platform_name=='B2C-PDD':
        dic1=pd.DataFrame({
            'screen_word':['CHANNEL'],
            'value_word':['Total'],
            'column_word':['CHANNEL'],
            'column_value':[['Total']]
        })

        dic2=pd.DataFrame({
            'screen_word':['CHANNEL','SEGMENT','SEGMENT','SEGMENT'],
            'value_word':['Total','三合一','冷食','热食'],
            'column_word':['CHANNEL','子品类','子品类','子品类'],
            'column_value':[['Total'],['三合一'],['冷食','Base Granola','Cold Others','Fancy Granola'],['热食','Base Hot','Hot Others','VA Hot']]
        })

        input_name='SEGMENT'

        channel_type_value=platform_polymerization_2(df_input,p_all_month,dic2,count_word,rule)

        channel_type_value1 = channel_type_value.iloc[1:].reset_index(drop=True)
        channel_type_value2  = channel_type_value.head(1)
        channel_type_value=pd.concat([channel_type_value1,channel_type_value2]).reset_index(drop=True)

    else:
        dic1=pd.DataFrame({
            'screen_word':['CHANNEL'],
            'value_word':['Total'],
            'column_word':['CHANNEL'],
            'column_value':[['Total']]
        })

        dic2=pd.DataFrame({
            'screen_word':['CHANNEL'],
            'value_word':['Total'],
            'column_word':['CHANNEL_TYPE_NEW'],
            'column_value':[column_list]
        })

        input_name='CHANNEL_TYPE_NEW'

        channel_type_value=platform_polymerization_2(df_input,p_all_month,dic2,count_word,rule)

    total_value=platform_polymerization_2(df_input,p_all_month,dic1,count_word,rule)


    if rule=='sale':
        value1=sale_count(total_value,column_pm_ty,column_pm_ly,column_ytd)
        value2=sale_count(channel_type_value,column_pm_ty,column_pm_ly,column_ytd)

    elif rule=='share':
        value1=share_count(total_value,column_pm_ty,column_pm_ly,column_ytd)
        value2=share_count(channel_type_value,column_pm_ty,column_pm_ly,column_ytd)


    df_list1.append(value1)
    df_list1.append(value2)


    for channel_type in channel_type_list:
        if channel_type=='Total':
            df_platform=df_total.copy()
        else:
            df_platform=df_total[df_total[input_name]==channel_type].reset_index(drop=True)

        df_input=df_platform.copy()

        dic_brand=pd.DataFrame({
            'screen_word':['brand'],
            'value_word':['Total'],
            'column_word':['brand'],
            'column_value':[brand_list+['Others','Total']]
        })
        df_brand=platform_polymerization_2(df_input,p_all_month,dic_brand,count_word,rule)
        
        if rule=='sale':
            brand_value=sale_count(df_brand,column_pm_ty,column_pm_ly,column_ytd)
        elif rule=='share':
            brand_value=share_count(df_brand,column_pm_ty,column_pm_ly,column_ytd)
        df_list2.append(brand_value)

    return df_list1,df_list2




#excel写入
def sheet_by_brand(df_total,wb,rule,name,sheet_name):

    value1=polymerization_total(df_total,'Value',rule,name)
    Volume1=polymerization_total(df_total,'Volume',rule,name)

    brand_value=polymerization_brand(df_total,'Value',rule)      
    brand_Volume=polymerization_brand(df_total,'Volume',rule)      

    start_col=4
    start_row=2
    ws = wb[sheet_name]
    for df in  tqdm([value1]+brand_value):
        df.set_index(['platform','Category/subcategory'], inplace=True)
        df.fillna('-',inplace=True)
        for r in dataframe_to_rows(df, index=False,header=False):
            for col_idx, cell_value in enumerate(r, start_col):
                ws.cell(row=start_row, column=col_idx, value=cell_value)
            start_row += 1
        start_row += 1

    start_col=3+len(df.columns)+2+1
    start_row=2
    for df in  tqdm([Volume1]+brand_Volume):
        df.set_index(['platform','Category/subcategory'], inplace=True)
        df.fillna('-',inplace=True)
        for r in dataframe_to_rows(df, index=False,header=False):
            for col_idx, cell_value in enumerate(r, start_col):
                ws.cell(row=start_row, column=col_idx, value=cell_value)
            start_row += 1
        start_row += 1



def sheet_by_platform(df_total,wb,sheet_name,column_list,channel_type_list,rule,platform_name,p_all_month):

    Value_list1,Value_list2=total_channel(df_total,'Value',rule,platform_name,column_list,channel_type_list,p_all_month)

    Volume_list1,Volume_list2=total_channel(df_total,'Volume',rule,platform_name,column_list,channel_type_list,p_all_month)

    if platform_name=='B2C-PDD':
        num1=3
        num2=3
    else:
        num1=2
        num2=1




    start_col=2
    start_row=2
    ws = wb[sheet_name]
    for df in  tqdm(Value_list1):
        df.set_index(['platform','Category/subcategory'], inplace=True)
        df.fillna('-',inplace=True)
        for r in dataframe_to_rows(df, index=False,header=False):
            for col_idx, cell_value in enumerate(r, start_col):
                ws.cell(row=start_row, column=col_idx, value=cell_value)
            start_row += 1
        start_row += num1
    
    for df in  tqdm(Value_list2):
        df.set_index(['platform','Category/subcategory'], inplace=True)
        df.fillna('-',inplace=True)
        for r in dataframe_to_rows(df, index=False,header=False):
            for col_idx, cell_value in enumerate(r, start_col):
                ws.cell(row=start_row, column=col_idx, value=cell_value)
            start_row += 1
        start_row += num2
    
    start_col=start_col+len(df.columns)+2+1
    start_row=2
    for df in  tqdm(Volume_list1):
        df.set_index(['platform','Category/subcategory'], inplace=True)
        df.fillna('-',inplace=True)
        for r in dataframe_to_rows(df, index=False,header=False):
            for col_idx, cell_value in enumerate(r, start_col):
                ws.cell(row=start_row, column=col_idx, value=cell_value)
            start_row += 1
        start_row += num1

    for df in  tqdm(Volume_list2):
        df.set_index(['platform','Category/subcategory'], inplace=True)
        df.fillna('-',inplace=True)
        for r in dataframe_to_rows(df, index=False,header=False):
            for col_idx, cell_value in enumerate(r, start_col):
                ws.cell(row=start_row, column=col_idx, value=cell_value)
            start_row += 1
        start_row += num2









def main (time):
    python_path,project_path,template_path,result_path=common.path()

    start  = datetime.now()
    ytd_list,year_list,pm_ty,pm_ly,p_24_month,p_17_month,p_18_month,column_ytd,column_year,column_pm_ty,column_pm_ly=time_total(time)

    brand_list=['桂格','卡乐比','西麦','家乐氏','卡芙莉','欣善怡','杂粮先生','五谷磨房','欧扎克','阿诺农场','福事多','雀巢','澳美佳','益汇坊','智力',
                '扎希菓乐','兵王的炊事班','皇麦世家','金味','燕谷坊','优品康','so acai','贝养颂','南国','捷森','三特','亨利','finax','精力沛','雅士利',
                '沃格尔','黑牛','皇室','鲍勃红磨坊','贝氏','老金磨方','三只松鼠','五谷食尚','鹊尚兰亭','润涵优品','金日阳','王饱饱','维地','麦粒美',
                '阿华田','好麦多','美味地带','家雀','旺旺','好哩','百草味','谷邻果乐','穗格氏','金日禾野','斯谷','固本堂','众德食品','三间磨坊','coles',
                '颜茴记','阳光蜜码','碧翠园','良品铺子','麦朵粒','维多麦','控碗','桑克拉','野山荞','蜜思诺','每麦烘焙','澳沃氏','纪米客','蒙北',
                '红色拖拉机','麸穗儿','澳菲顿','世壮','罗生记','德富祥','田园主义','呵点儿','禾力康','可益康','山姆','彩虹星球','燕之坊','窝小芽',
                '小黄象','轻食兽','捷氏','初萃','阴山优麦','红锚','第三主粮','荷兰乳牛','智仁','好想你','山萃','周氏','秦老太']

    df_total=sql_select(brand_list)


    current_time = datetime.now()
    formatted_time = current_time.strftime("%Y-%m-%d %H_%M_%S")

    template=f'{template_path}/桂格麦片模板_2024.xlsx'
    filename = f'{result_path}/桂格麦片_{formatted_time}.xlsx'
    wb = openpyxl.load_workbook(template)

    sheet_by_brand(df_total,wb,'sale','Total Cereal','Sales by brand (Total Cereal)')

    sheet_by_brand(df_total,wb,'share','Total Cereal','Share by brand (Total Cereal)')

    df_input=df_total[df_total['CHANNEL'].isin(['B2C-Tmall','B2C-standalone B2C'])].reset_index(drop=True)
    sheet_by_brand(df_input,wb,'sale','EC',' Sales by brand EC')

    sheet_by_brand(df_input,wb,'share','EC','Share by brand EC')





    column_list=['Tmall Flagship','Tmall Supermarket','Tmall B store','Tmall Overseas','Total']
    channel_type_list=['Total','Tmall Flagship','Tmall Supermarket','Tmall B store','Tmall Overseas']

    df_input=df_total[df_total['CHANNEL']=='B2C-Tmall'].reset_index(drop=True)


    sheet_by_platform(df_input,wb,'Sales by Tmall',column_list,channel_type_list,'sale','B2C-Tmall',p_17_month)
    
    sheet_by_platform(df_input,wb,'Share by Tmall',column_list,channel_type_list,'share','B2C-Tmall',p_17_month)




    column_list=['Standalone Overseas','Standalone pop platform','Standalone self run','Total']
    channel_type_list=['Total','Standalone Overseas','Standalone pop platform','Standalone self run']

    df_input=df_total[df_total['CHANNEL']=='B2C-standalone B2C'].reset_index(drop=True)

    sheet_by_platform(df_input,wb,'Sales by Standalone B2C',column_list,channel_type_list,'sale','B2C-standalone B2C',p_17_month)

    sheet_by_platform(df_input,wb,'Share by Standalone B2C',column_list,channel_type_list,'share','B2C-standalone B2C',p_17_month)


    
    df_input=df_total[df_total['CHANNEL']=='B2C-PDD'].reset_index(drop=True)

    channel_type_list=['Total','冷食','三合一','热食']

    sheet_by_platform(df_input,wb,'Sales by PDD',column_list,channel_type_list,'sale','B2C-PDD',p_18_month)

    sheet_by_platform(df_input,wb,'Share by PDD',column_list,channel_type_list,'share','B2C-PDD',p_18_month)


    wb.save(filename)

    end  = datetime.now()
    print("程序运行时间："+str((end-start).seconds)+"秒")

if __name__ =="__main__":

    received_arguments = sys.argv[1:]
    time_begin=received_arguments[0]
    time_end=received_arguments[1]
    main(time_begin)
